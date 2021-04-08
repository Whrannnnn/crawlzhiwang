import re
from bs4 import BeautifulSoup
import requests
from requests import RequestException
import openpyxl
import time
from selenium import webdriver

cnt = 2


def get_page(url):
    try:
        # 添加User-Agent，放在headers中，伪装成浏览器
        headers = {
                'user-agent': 'xx',
                'cookie': 'xx'
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            response.encoding = response.apparent_encoding
            return response.text
        return None
    except RequestException as e:
        print(e)
        return None


def get_info(page, html):
    # 1: 创建webdriver驱动：
    driver = webdriver.Chrome()
    # 2: 发送url请求
    driver.get(html)
    # 3: 操作页面：

    title_list = []
    author_list = []
    keyword_list = []
    url_list = []

    for i in range(page):
        driver.find_element_by_xpath('//*[@id="PageContent"]/div[1]/div[2]/div[13]/a[11]').click()

        # ****************写入url******************
        # xpath定位
        # driver.find_elements_by_xpath和 driver.find_element_by_xpath的区别:
        # elements爬取符合条件的所有数据并返回列表,找不到返回空列表
        # element爬取符合条件的一条数据,如果找不到报错
        time.sleep(2.5)
        content = driver.find_elements_by_xpath('//*[@id="article_result"]/div/div/p[1]/a[1]')
        # 循环列表 取出每个selenium对象
        for item in content:
            # 取出标题的href链接并打印
            # print(item.text)
            # # print(item.text)
            # print(len(item.text))
            # title = item.text.split(' ')[0]
            # print(title)
            url_list.append(item.get_attribute('href'))
            title_list.append(item.text.split(' ')[0])
        #  ****************************************
        time.sleep(2)
        # ****************写入author***************
        authors = driver.find_elements_by_xpath('//*[@id="article_result"]/div/div/p[3]/span[1]')
        for author in authors:
            author_list.append(author.get_attribute('title'))
        #  *****************************************
        time.sleep(2)
        # *******************写入key_word***************
        key_words = driver.find_elements_by_xpath('//*[@id="article_result"]/div/div/div[1]/p[1]/a[1]')
        for keyword in key_words:
            keyword_list.append(keyword.get_attribute('data-key'))
        #  ********************************************
        print("**********第{}页爬取完成**********".format(i + 1))
        time.sleep(2)
    driver.quit()
    return title_list, author_list, keyword_list, url_list


def get_abstract(url_list):
    abstract = []
    cnt = 2
    for url in url_list:
        # entity_url = 'http://' + url
        try:
            # 添加User-Agent，放在headers中，伪装成浏览器
            kv = {
                    'user-agent': 'xx',
                    'cookie': 'xx'
            }
            r = requests.get(url, headers=kv, timeout=30)
            r.raise_for_status()
            r.encoding = r.apparent_encoding
            demo = r.text
            soup = BeautifulSoup(demo, 'lxml')
            rst = soup.select('div.xx_font')[0]
            # 将类型为<class 'bs4.element.ResultSet'>的rst转换为str类型变量
            soup_str = str(rst)
            pattern = re.compile('</font>([\u4e00-\u9fa5-a-zA-Z0-9\;\"\!\！\`\·\[\]\【\】\｜\@\*\^\$\—\“\”\、\.\,\，\。\;\；\:\ \：\《\》\/\（\）\(\)\s+]+)',
                                 re.S)
            abstract = pattern.findall(soup_str)
            if len(abstract) == 0:
                print("NULL")
                xfile = openpyxl.load_workbook('../test.xlsx')
                new_sheet = xfile.worksheets[0]
                new_sheet.cell(cnt, 5).value = "None"
                xfile.save('../test.xlsx')
                cnt += 1
                print('**********第{}次写入成功**********'.format(cnt - 2))
            else:
                # 写入摘要
                xfile = openpyxl.load_workbook('../test.xlsx')
                new_sheet = xfile.worksheets[0]
                new_sheet.cell(cnt, 5).value = abstract[0]
                xfile.save('../test.xlsxx')
                cnt += 1
                print('**********第{}次写入成功**********'.format(cnt - 2))
        except:
            try:
                if len(abstract[0]) > 0:
                    xfile = openpyxl.load_workbook('../test.xlsx')
                    new_sheet = xfile.worksheets[0]
                    new_sheet.cell(cnt, 5).value = abstract[0]
                    cnt += 1
                    xfile.save('../test.xlsx')
                    print('**********第{}次写入成功**********'.format(cnt - 2))
                else:
                    print("NULL")
                    xfile = openpyxl.load_workbook('../test.xlsx')
                    new_sheet = xfile.worksheets[0]
                    new_sheet.cell(cnt, 5).value = "None"
                    cnt += 1
                    xfile.save('../test.xlsx')
                    print('**********第{}次写入成功**********'.format(cnt - 2))
            except:
                print("NULL")
                xfile = openpyxl.load_workbook('../test.xlsx')
                new_sheet = xfile.worksheets[0]
                new_sheet.cell(cnt, 5).value = "None"
                cnt += 1
                xfile.save('../test.xlsx')
                print('**********第{}次写入成功**********'.format(cnt - 2))


def data_write(file_path, datas):
    # ***********************************通过openyxl进行Excel的写入，没发现问题***************************************
    wb = openpyxl.Workbook()
    ws = wb.active  # 默认插在最后
    ws.title = 'demo_sheet1'
    ws['A1'] = '题目'
    ws['B1'] = "作者"
    ws['C1'] = "关键词"
    ws['D1'] = "url链接"
    ws['E1'] = "摘要"
    # 将数据写入第 i 行，第 j 列
    column = 1
    for i in datas.keys():
        for j in range(len(datas[i])):
            ws.cell(j + 2, column).value = datas[i][j]
        column = column + 1
    wb.save('/test.xlsx')
    print('**********前部分数据写入成功**********')


if __name__ == '__main__':
    headers = {
        'user-agent': 'xx',
        'cookie': 'xx'
    }
    # 可以交互输入 也可以直接指定
    key_word = input('请输入搜索关键词：')
    page = int(input('检索页数：'))
    # 从哪一页开始爬 爬几页
    # start_page = int(input('请输入爬取的起始页：'))
    base_url = 'http://search.cnki.com.cn/Search/Result?content={}'
    first_url = base_url.format(key_word)
    htm1 = requests.get(first_url, headers=headers)
    soup = BeautifulSoup(htm1.text, 'html.parser')
    print(first_url)
    title_list, author_list, key_data_list, url_list = get_info(page, first_url)

    output = {'title': title_list, 'author': author_list, 'keyword': key_data_list, 'url': url_list}

    data_write('data1.xls', output)

    # 获取每篇论文的摘要
    get_abstract(url_list)
    time.sleep(2)  # 间隔2s
